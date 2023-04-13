import openpyxl

def read_excel(file_name: str, worksheet_name: str):
    """
    Fungsi untuk membaca data dari file Ms. Excel/XLSX dan mengonversinya menjadi list of dictionary
    args:
        - file_name (str): nama file excel
        - worksheet_name (str): nama worksheet tempat data berada
    return:
        - data (list): list of dictionary dari data yang telah dikonversi
    """
    
    workbook = openpyxl.load_workbook(file_name)
    worksheet = workbook[worksheet_name]

    data = []
    headers = []

    for item in range(1, worksheet.max_column + 1):
        headers.append(worksheet.cell(row=1, column=item).value)

    for row in worksheet.iter_rows(min_row=2, values_only=True):
        row_data = {}

        for item in range(len(headers)):
            key = headers[item]
            value = row[item]

            row_data[key] = value

        data.append(row_data)

    return data