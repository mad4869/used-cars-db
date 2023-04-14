import openpyxl

def read_excel(file_name: str, worksheet_name: str):
    """
    Fungsi untuk membaca data dari file Excel/XLSX dan mengonversinya menjadi list of dictionary
    args:
        - file_name (str): nama file excel
        - worksheet_name (str): nama worksheet tempat data berada
    return:
        - data (list): list of dictionary dari data yang telah dikonversi
    """
    # Mengakses workbook dan worksheet dari file excel
    workbook = openpyxl.load_workbook(file_name)
    worksheet = workbook[worksheet_name]

    # Inisialisasi list data dan headers
    data = []
    headers = []

    # Looping sebanyak kolom worksheet untuk memasukkan nama kolom ke list headers
    for item in range(1, worksheet.max_column + 1):
        headers.append(worksheet.cell(row=1, column=item).value)

    # Looping sebanyak baris worksheet untuk memasukkan data sesuai kolomnya
    for row in worksheet.iter_rows(min_row=2, values_only=True):
        row_data = {}

        # Memasangkan key and value item sesuai kolom
        for item in range(len(headers)):
            key = headers[item]
            value = row[item]

            row_data[key] = value

        # Memasukkan dictionary ke list data
        data.append(row_data)

    return data